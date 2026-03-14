from django.core.management.base import BaseCommand, CommandError
from user_record.models import UserRecord
import openpyxl


class Command(BaseCommand):
    help = "Imports specified xlsx file into User Record objects"

    def add_arguments(self, parser):
        parser.add_argument("file")

    def handle(self, *args, **options):
        # Load specified file, throw CommandError if not found
        try:
            df = openpyxl.load_workbook(options["file"])
            df1 = df.active
        except:
            raise CommandError('Unable to open file {}'.format(options["file"]))

        # Reads data rows, zips to UserRecord fields and creates UserRecords
        fields = ["first_name", "last_name", "job_title", "phone", "email"]
        for row in range(1, df1.max_row):
            vals = []
            for col in df1.iter_cols(1, df1.max_column):
                vals.append(col[row].value)
            u = dict(zip(fields ,vals))
            # Passes if UserRecord creation errors, so subsequent records created
            try:
                UserRecord.objects.create(**u)
            except:
                pass

        self.stdout.write(
            self.style.SUCCESS('User records successfully imported from {}'.format(options["file"]))
        )